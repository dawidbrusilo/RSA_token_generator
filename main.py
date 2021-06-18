import json
import time
from configparser import ConfigParser
from selenium import webdriver
import openpyxl
import os
import sys
import PySimpleGUI as sg


def resource_path(relative_path: str) -> str:
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.dirname(__file__)
    return os.path.join(base_path, relative_path)


def gui():
    sg.theme('dark')
    layout = [
        [sg.Text('Username')],
        [sg.InputText(size=(25, 1), key='username')],
        [sg.Text('Pin+token')],
        [sg.InputText(password_char='*', key='password', size=(25, 1))],
        [sg.Text('Number of users')],
        [sg.InputText(size=(5, 1), key='number_of_users')],
        [sg.Button("Run script", button_color='#858585')]]

    # Create the window
    window = sg.Window("RSA Automated", layout, size=(250, 200))

    # Create an event loop
    while True:
        event, values = window.read()
        # End program if user closes window or
        # presses the OK button
        if event == "Run script":
            RsaUsername = values['username']
            RsaPassword = values['password']
            # Number_of_users = int(values['number_of_users'])

            if not RsaUsername:
                sg.Popup('Opps', 'Username field is empty!')
            if not RsaPassword:
                sg.Popup('Opps', 'Password field is empty!')
            try:
                Number_of_users = int(values['number_of_users'])
            except ValueError:
                sg.Popup('Opps', 'Number of users field is empty or value is a string!')

            else:

                return RsaUsername, RsaPassword, Number_of_users

        else:
            event == sg.WIN_CLOSED
            sys.exit()
            break

    window.close()


def main():
    RsaUsername, RsaPassword, Number_of_users = gui()
    Number_of_users += 2
    config = ConfigParser()
    config.read("config.ini")
    CHROME_DRIVER_PATH = config.get("chromedriver", "path")
    DURATION = config.getint("delay", "seconds")
    USERS = config.getint("number_of_users", "users")
    users = USERS + 2
    URL = config.get("website", "url")

    excel = openpyxl.load_workbook("data\Rsa.xlsx")
    sheet = excel.active
    driver = webdriver.Chrome(resource_path(CHROME_DRIVER_PATH))
    driver.get(URL)
    driver.implicitly_wait(30)
    user_id = driver.find_element_by_name('userName')
    user_id.send_keys(RsaUsername)
    password1 = driver.find_element_by_name('passCode')
    password1.send_keys(RsaPassword)
    driver.implicitly_wait(30)
    driver.find_element_by_id('login').click()
    for x in range(2, Number_of_users):
        usernameExcel = sheet.cell(row=x, column=1).value
        driver.implicitly_wait(30)
        search_box = driver.find_element_by_id('tsearch')
        search_box.send_keys(usernameExcel)
        driver.implicitly_wait(30)
        driver.find_element_by_id('userview').click()
        driver.implicitly_wait(30)
        driver.find_element_by_id('userArrow').click()
        driver.implicitly_wait(30)
        driver.find_element_by_id('ui-id-16').click()
        driver.implicitly_wait(30)
        driver.find_element_by_id('assignTokenSubmit').click()
        driver.implicitly_wait(30)
        driver.find_element_by_id('tokenArrow').click()
        driver.implicitly_wait(30)
        driver.find_element_by_id('ui-id-23').click()
        driver.implicitly_wait(30)
        driver.find_element_by_id('emergencyAccessfixed').click()
        time.sleep(DURATION)
        driver.implicitly_wait(30)
        token = driver.find_element_by_xpath("//span[@id='emergencyTokenCode']").text
        sheet.cell(row=x, column=2).value = token
        excel.save("data\Rsa.xlsx")
        driver.refresh()

    driver.quit()


if __name__ == "__main__":
    main()
